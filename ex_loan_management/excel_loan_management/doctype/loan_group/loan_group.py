# Copyright (c) 2025, Rutuja Somvanshi and contributors
# For license information, please see license.txt

# import frappe
from frappe.model.document import Document
from openpyxl import load_workbook
import frappe
from frappe.utils.file_manager import save_file
from ex_loan_management.api.utils import api_error
import re
from datetime import datetime
class LoanGroup(Document):
    def autoname(self):
        prefix = "GC"

        # Fetch latest member_id (not name)
        last_record = frappe.db.get_all(
            "Loan Group",
            filters={"group_id": ["like", f"{prefix}%"]},
            fields=["group_id"],
            order_by="group_id desc",
            limit=1
        )

        if last_record:
            last_id = last_record[0].group_id
            match = re.search(r"GC(\d+)", last_id)
            new_num = int(match.group(1)) + 1 if match else 1
        else:
            new_num = 1
        # Safe padding
        if new_num <= 9999:
            new_id = f"{prefix}{new_num:04d}"
        else:
            new_id = f"{prefix}{new_num}"
        self.name = new_id
        self.group_id = new_id

    
    def on_update(self):
        # When workflow is approved, create loan group assignment
        if self.workflow_state == "Approved":
            create_loan_group_assigned(self)


def create_loan_group_assigned(doc):
    """Create a Loan Group Assignment record when Loan Group is approved"""

    # Step 1: Check if group is already assigned to someone
    existing_assignment = frappe.db.get_value(
        "Loan Group Assignment", 
        {"loan_group": doc.name}, 
        ["employee"]
    )
    print("existing_assignment ....",existing_assignment)

    if existing_assignment:
        return  # stop, since group already assigned

    # Step 2: Get Employee linked with the Loan Group owner (via user_id)
    employee = frappe.db.get_value("Employee", {"user_id": doc.owner}, "name")
    if not employee:
        return

    # Step 3: Create new Loan Group Assignment record
    assigned = frappe.new_doc("Loan Group Assignment")
    assigned.loan_group = doc.name
    assigned.employee = employee
    assigned.start_date = datetime.now()
    assigned.owner = doc.owner  # make owner same as the user who owns the Loan Group
    assigned.insert(ignore_permissions=True)

    frappe.msgprint(f"✅ Loan Group '{doc.name}' assigned successfully to Employee '{employee}'.")



@frappe.whitelist()
def import_loan_groups(file_url):
    from frappe.utils.file_manager import get_file

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    filename = file_doc.file_name.lower()

    rows = []
    print("file_doc ....", file_doc, filename)

    # Read Excel file and convert rows to dictionary
    if filename.endswith(".xlsx"):
        file_path = file_doc.get_full_path()
        with open(file_path, "rb") as f:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]  # First row as headers
            print("headers ....", headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

    skipped = []
    updated_groups = []
    

    for row in rows:
        member_id = str(row.get("MEMBER NO")).strip() if row.get("MEMBER NO") else None
        if not member_id:
            continue  # Skip if no member_id

        group_id = str(row.get("GROUP CODE")).strip() if row.get("GROUP CODE") else None
        if not group_id:
            continue  # Skip if no group_id
        head_name = row.get("NAME OF GROUP HEAD")
        head_mobile_no = row.get("GROUP HEAD MOB NO")
        print('head_name ......',head_name,"head_mobile_no ......",head_mobile_no)

        # Check if Loan Group exists
        group_name = frappe.db.exists("Loan Group", {"group_id": group_id})
        if group_name:
            # Check if Loan Member exists
            member_name = frappe.db.exists("Loan Member", {"member_name": head_name,"mobile_no":head_mobile_no})
            print('member_name ......',member_name)
            if member_name:
                # Load Loan Group and update loan_head
                group_doc = frappe.get_doc("Loan Group", group_name)
                group_doc.group_head = member_name  # loan_head is a Link field to Loan Member
                group_doc.save(ignore_permissions=True)
                updated_groups.append(group_name)
            else:
                skipped.append(member_id)
        else:
            skipped.append(f"Group:{group_id}")

    frappe.db.commit()

    # Prepare Summary
    msg = f"✅ Processed {len(rows)} rows."
    if updated_groups:
        msg += f"\n🔗 Updated {len(updated_groups)} Loan Groups with loan_head: {', '.join(updated_groups)}"
    if skipped:
        msg += f"\n⏩ Skipped {len(skipped)} items: {', '.join(skipped)}"

    return msg



import frappe
from ex_loan_management.api.utils import get_paginated_data

"""
    Get Loan Group List (with optional pagination, search & sorting)
"""
@frappe.whitelist()
def loan_group_list(page=1, page_size=10,sort_by="group_name", sort_order="asc", search=None, is_pagination=False):
    extra_params = {"search": search} if search else {}
    base_url = frappe.request.host_url.rstrip("/") + frappe.request.path

    user = frappe.session.user
    filters = {}
    if "Agent" in frappe.get_roles(user):
        print("in if ....")
        # Get employee id
        employee_id = frappe.db.get_value("Employee", {"user_id": user}, "name")
        if not employee_id:
            return []

        # Get loan groups for this employee
        groups = frappe.get_all(
            "Loan Group Assignment",
            filters={"employee": employee_id},
            pluck="loan_group"
        )
        
        if groups:
            filters["name"] = ["in", groups]

        else:
            # Members without a group assigned (group is null/empty)
            filters["name"] = ["=", ""]

    return get_paginated_data(
        doctype="Loan Group",
        fields=["name", "group_name", "group_head","group_image"],
        search=search,
        filters=filters,
        sort_by=sort_by,
        sort_order=sort_order,
        page=int(page),
        page_size=int(page_size),
        search_fields=["group_name", "group_head","group_image"],
        is_pagination=frappe.utils.sbool(is_pagination),
        base_url=base_url,
        extra_params=extra_params,
        link_fields={"group_head": "member_name"},
        image_fields=["group_image"]
    )




@frappe.whitelist()
def create_loan_group():
    try:
        data = frappe.form_dict  # works for JSON body and form-data
        files = frappe.request.files  # uploaded files

        #
        doc = frappe.get_doc({
            "doctype": "Loan Group",
            "group_id":data.get("group_id"),
            "group_name":data.get("group_name"),
            "group_head":data.get("group_head"),
        })
        doc.insert(ignore_permissions=True)
        
        if "group_image" in files:
            upload = files["group_image"] 
            print("upload ...",upload,'... ', upload.filename)
            if upload or upload.filename:
                print("in if ....")
                file_doc = save_file(
                    fname=upload.filename,
                    content=upload.stream.read(),
                    dt="Loan Group",
                    dn=doc.name,
                    is_private=0
                )
                doc.set("group_image", file_doc.file_url)


        # Step 2: Insert Collection In Hand (runs validate() automatically)
        doc.save(ignore_permissions=True)
        doc.db_set("workflow_state", "Pending", update_modified=False)
        frappe.db.commit()
        

        return {
            "status": "success",
            "status_code": 201,
            "msg": "Loan Group Created Successfully"
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Loan Group API Error")
        return api_error(e)





@frappe.whitelist()
def update_loan_group(name):
    try:
        data = frappe.form_dict  # works for JSON body and form-data
        files = frappe.request.files  # uploaded files
        
        loan_group_id = name
        if not loan_group_id:
            return {"status": "error", "message": "Loan Group ID is required"}

        # # Fetch existing loan member doc
        doc = frappe.get_doc("Loan Group", loan_group_id)

        if doc.workflow_state == "Pending":
            frappe.throw("Cannot update Loan Group in Pending state.")
        
        if data.get("group_name"):
            doc.group_name = data.get("group_name")
        if data.get("group_head"):
            doc.group_head = data.get("group_head")


        if "group_image" in files:
            upload = files["group_image"] 
            if upload or upload.filename:
                file_doc = save_file(
                    fname=upload.filename,
                    content=upload.stream.read(),
                    dt="Loan Group",
                    dn=doc.name,
                    is_private=0
                )
                doc.set("group_image", file_doc.file_url)


        # Step 2: Insert Collection In Hand (runs validate() automatically)
        doc.save(ignore_permissions=True)
        doc.db_set("workflow_state", "Pending", update_modified=False)

        frappe.db.commit()

        return {
            "status": "success",
            "status_code": 201,
            "msg": "Loan Group Updated Successfully"
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Loan Group API Error")
        return api_error(e)


