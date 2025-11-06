# Copyright (c) 2025, Rutuja Somvanshi and contributors
# For license information, please see license.txt

from frappe.model.document import Document
import frappe
from openpyxl import load_workbook
from datetime import datetime
from frappe.utils.file_manager import save_file, get_file
from ex_loan_management.api.utils import get_paginated_data, api_error
from urllib.parse import urljoin
import re
from datetime import date

class LoanMember(Document):

    def autoname(self):
        prefix = "TMA"

        # Fetch latest member_id (not name)
        last_record = frappe.db.get_all(
            "Loan Member",
            filters={"member_id": ["like", f"{prefix}%"]},
            fields=["member_id"],
            order_by="member_id desc",
            limit=1
        )

        if last_record:
            last_id = last_record[0].member_id
            match = re.search(r"TMA(\d+)", last_id)
            new_num = int(match.group(1)) + 1 if match else 1
        else:
            new_num = 1
        # Safe padding
        if new_num <= 99999:
            new_id = f"{prefix}{new_num:05d}"
        else:
            new_id = f"{prefix}{new_num}"
        self.name = new_id
        self.member_id = new_id

    def before_save(self):
        # ✅ Skip validation if called during import
        if getattr(frappe.flags, "in_import", False):
            return

        self.completed_age, self.entry_age= calculate_member_age(self.dob)

        prev_status = self.get_db_value("status") if not self.is_new() else None

        if self.status =="Pending":
            # Fetch all fieldnames of this DocType except system/internal fields
            docfields = [df.fieldname for df in frappe.get_meta(self.doctype).fields 
                         if df.fieldtype not in ["Section Break", "Column Break", "Tab Break", "Table"]]

            # Check each field is filled
            missing_fields = []
            for fieldname in docfields:
                value = self.get(fieldname)
                if fieldname not in ["group","aadhar_image_back", "pancard_image_back","voter_id_image_back",
                                     "address_line_2","cibil_score","bank_address","mobile_no_2",
                                    "cibil_date"]:
                    if value in [None, ""]:
                        missing_fields.append(fieldname)

            if missing_fields:
                frappe.throw(f"To set status as 'Pending', the following fields must be filled: {', '.join(missing_fields)}")

        if self.status != "Pending":
            print(self.address_verified , self.pancard_verified , self.aadhar_verified , self.voter_id_verified)
            if prev_status == "Draft" and self.status != "Draft":
                frappe.throw("Status can only be changed from 'Draft' to 'Pending'.")
            if self.address_verified and self.pancard_verified and self.aadhar_verified and self.voter_id_verified:
                self.status = "Verified"
            if self.status == "Verified" and (not self.address_verified or not self.pancard_verified or not self.aadhar_verified or not self.voter_id_verified):
                frappe.throw("To set status as 'Verified', all verifications must be completed.")
        
        self.member_name = f"{self.first_name or ''} {self.middle_name or ''} {self.last_name or ''}".strip()

        existing_member = check_unique_member(self.member_name, self.pancard, self.name)
        if existing_member:
            frappe.throw(existing_member)

        if not self.created_by:
            # Get the full name from the User DocType
            owner_full_name = frappe.db.get_value("User", self.owner, "full_name")
            self.db_set("created_by",owner_full_name)


    def validate(self):
        # ✅ Validate Voter ID
        if self.voter_id:
            voter_pattern = r'^[A-Z]{3}[0-9]{7}$'
            if not re.match(voter_pattern, self.voter_id):
                frappe.throw("Invalid Voter ID format. It should be like 'ABC1234567' (3 letters followed by 7 digits).")

        # ✅ Validate PAN Card
        if self.pancard:
            pan_pattern = r'^[A-Z]{5}[0-9]{4}[A-Z]{1}$'
            if not re.match(pan_pattern, self.pancard):
                frappe.throw("Invalid PAN Card format. It should be like 'ABCDE1234F' (5 letters, 4 digits, 1 letter).")

         # Validate both mobile numbers
        self.validate_mobile_no(self.mobile_no, fieldname="Mobile No",  required=True)
        self.validate_mobile_no(self.mobile_no_2, fieldname="Alternate Mobile No", required=False)

    def validate_mobile_no(self, number, fieldname="Mobile No", required=True):
        # Extract digits after +91
        print("number ....",number , ' fieldname ....',fieldname)
        digits = number.replace("+91", "").strip()
        if not digits:
            if required:
                frappe.throw(f"{fieldname} is required")
            else:
                return  # optional blank field is allowed

        # Ensure +91 prefix
        if number:
            if not number.startswith("+91"):
                frappe.throw(f"{fieldname} must start with +91")
            
        # Check that exactly 10 digits remain
        if not digits.isdigit() or len(digits) != 10:
            frappe.throw(f"Enter a valid 10-digit number after +91 for {fieldname}")



@frappe.whitelist()
def import_loan_members(file_url):
    frappe.flags.in_import = True  # ✅ Set import flag
    try:
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        filename = file_doc.file_name.lower()
        rows = []

        # Read Excel file and convert rows to dictionary
        if filename.endswith(".xlsx"):
            file_path = file_doc.get_full_path()
            with open(file_path, "rb") as f:
                wb = load_workbook(f, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]  # First row as headers
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

        skipped = []
        created_members = []
        created_groups = []
        created_branches = []
        created_occupations = []

        for row in rows:
            print("row  ....",row)
            member_id = str(row.get("MEMBER NO")).strip() if row.get("MEMBER NO") else None
            if not member_id:
                continue  # Skip if no member_id

            # Member basic info
            member_name = row.get("NAME OF MEMBER")
            type_of_borrower = row.get("TYPE OF BORROWER")
            mobile_no = row.get("MOB NO")
            gender = row.get("GENDER")
            email = ""
            dob = row.get("DOB")
            entry_age = row.get("ENTRY AGE")
            completed_age = row.get("COMPLETED AGE")
            occupation = row.get("OCCUPATION")
            address = row.get("ADDRESS")
            state = row.get("STATE")
            if str(state).upper() == "MAHARASHTRA":
                state = "MH"
            country = "India"
            city = row.get("CITY")
            pincode = row.get("PIN CODE")

            # Group and branch info
            group_id = str(row.get("GROUP CODE")).strip() if row.get("GROUP CODE") else None
            group_name = row.get("NAME OF GROUP")
            branch_code = str(row.get("BRANCH CODE")).strip() if row.get("BRANCH CODE") else None
            branch_name = row.get("BRANCH NAME")
            print("branch_name .....",branch_name, '  mobile_no ....',mobile_no)
            # 1️⃣ Create Loan Group if not exists
            if group_id and not frappe.db.exists("Loan Group", {"group_id": group_id}):
                grp_doc = frappe.new_doc("Loan Group")
                grp_doc.group_id = group_id
                grp_doc.group_name = group_name
                grp_doc.insert(ignore_permissions=True)
                created_groups.append(group_id)

            # 2️⃣ Create Branch if not exists
            if branch_code and not frappe.db.exists("Branch", {"branch": branch_code}):
                br_doc = frappe.new_doc("Branch")
                br_doc.branch = branch_code
                br_doc.branch_name = branch_name
                br_doc.insert(ignore_permissions=True)
                created_branches.append(branch_code)

            # 3️⃣ Create Occupation if not exists
            occupation_link = None
            if occupation:
                occupation = occupation.strip()
                if not frappe.db.exists("Occupation", {"occupation": occupation}):
                    occ_doc = frappe.new_doc("Occupation")
                    occ_doc.occupation = occupation
                    occ_doc.insert(ignore_permissions=True)
                    created_occupations.append(occupation)
                occupation_link = frappe.db.get_value("Occupation", {"occupation": occupation})

            # 4️⃣ Skip if Loan Member already exists
            if frappe.db.exists("Loan Member", {"member_id": member_id}):
                skipped.append(member_id)
                continue
            if type_of_borrower == 'BORROWER':
                type_of_borrower ='Borrower'
            else:
                type_of_borrower = 'Co-Borrower'
            
            first_name, middle_name, last_name = split_name(member_name)
                
            try:
                print('in try ........')
                # 5️⃣ Create Loan Member
                doc = frappe.new_doc("Loan Member")
                doc.member_id = member_id
                doc.member_name = member_name
                doc.first_name = first_name
                doc.middle_name = middle_name
                doc.last_name = last_name
                doc.status = "Verified"
                doc.type_of_borrower = type_of_borrower
                doc.mobile_no = f"+91{mobile_no}" if mobile_no else None
                doc.mobile_no_2 = ""
                doc.gender = gender
                doc.email = email
                doc.dob = dob
                doc.entry_age = entry_age
                doc.completed_age = completed_age
                doc.occupation = occupation_link
                doc.address = address
                doc.state = state
                doc.country = country
                doc.city = city
                doc.pincode = pincode
                doc.aadhar = row.get("AADHAR NO")
                doc.pancard = row.get("PAN")
                doc.address_doc_type= row.get("Electricity Bill")
                doc.aadhar_verified = True
                doc.pancard_verified = True
                doc.address_verified= True
                doc.voter_id_verified = True
                doc.bank_name = row.get("BANK NAME")
                doc.account_number = row.get("SAVING ACCOUNT NUMBER")
                doc.branch = row.get("BRANCH NAME")
                doc.bank_address = row.get("BRANCH NAME")
                doc.holder_name = member_name
                doc.ifsc_code = row.get("IFSC CODE")
                doc.account_type = "Saving"

                print('doc ...........', doc, '.....', member_id)

                # Link fields
                if group_id:
                    group_id=frappe.db.exists("Loan Group", {"group_id": group_id})
                    doc.group = group_id
                if branch_code:
                    branch_code = frappe.db.exists("Branch", {"branch_code": branch_code})
                    doc.branch_code = branch_code

                doc.insert(ignore_permissions=True)
                created_members.append(member_id)
            except Exception as e:
                print(f"Error creating Loan Member {member_id}: {e}")
        frappe.db.commit()

        # Prepare Summary
        msg = f"✅ Created {len(created_members)} Loan Members."
        if created_groups:
            msg += f"\n🏷 Created {len(created_groups)} Loan Groups: {', '.join(created_groups)}"
        if created_branches:
            msg += f"\n🏢 Created {len(created_branches)} Branches: {', '.join(created_branches)}"
        if created_occupations:
            msg += f"\n💼 Created {len(created_occupations)} Occupations: {', '.join(created_occupations)}"
        if skipped:
            msg += f"\n⏩ Skipped {len(skipped)} existing members: {', '.join(skipped)}"

        return msg
    finally:
        frappe.flags.in_import = False  # ✅ Always reset





def split_name(full_name):
    """Split full name into first, middle, last"""
    parts = full_name.strip().split()
    if len(parts) == 1:
        return parts[0], "", ""   # only first name
    elif len(parts) == 2:
        return parts[0], "", parts[1]   # first + last
    elif len(parts) == 3:
        return parts[0], parts[1], parts[2]  # first + middle + last
    else:
        # more than 3 → treat first as first_name, last as last_name, middle = everything in between
        return parts[0], " ".join(parts[1:-1]), parts[-1]


def check_unique_member(member_name, pancard, current_name=None):
    """Check if Loan Member with same name & PAN exists (excluding current)."""
    existing = frappe.db.exists(
        "Loan Member",
        {
            "member_name": member_name,
            "pancard": pancard,
            "name": ["!=", current_name] if current_name else ["!=", ""]
        }
    )

    if existing:
        return "Loan member with the same name and PAN already exists."
    return None

    
def calculate_member_age(dob):

    """
    dob: datetime.date or string in 'YYYY-MM-DD'
    Returns: dict with completed_age and current_age
    """
    if isinstance(dob, str):
        dob = date.fromisoformat(dob)

    today = date.today()

    # Completed age in full years
    completed_age = today.year - dob.year
    if (today.month, today.day) < (dob.month, dob.day):
        completed_age -= 1

    # Current age = completed + 1
    current_age = completed_age + 1

    return  completed_age, current_age


@frappe.whitelist()
def import_update_loan_members(file_url):
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    filename = file_doc.file_name.lower()

    rows = []

    # Read Excel file and convert rows to dictionary
    if filename.endswith(".xlsx"):
        file_path = file_doc.get_full_path()
        with open(file_path, "rb") as f:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]  # First row as headers

            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

    skipped = []
    created_members = []

    for row in rows:
        member_id = str(row.get("MEMBER NO")).strip() if row.get("MEMBER NO") else None
        if not member_id:
            continue  # Skip if no member_id

        # Member basic info
        member_name = row.get("NAME OF MEMBER")
        member_exists = frappe.db.exists("Loan Member", {"member_id": member_id})
        
        # If Loan Member already exists
        if member_exists:
            first_name, middle_name, last_name = split_name(member_name)
            
            # 5️⃣ Create Loan Member
            doc = frappe.get_doc("Loan Member", member_exists)  
            doc.member_name = member_name
            doc.first_name = first_name
            doc.middle_name = middle_name
            doc.last_name = last_name
            doc.status = "Verified"
            doc.aadhar_verified = True
            doc.pancard_verified = True
            doc.address_verified = True
            doc.company = "Excellminds (Demo)"
            doc.save(ignore_permissions=True)
            created_members.append(member_id)
            
    frappe.db.commit()

    # Prepare Summary
    msg = f"✅ Created {len(created_members)} Loan Members."
    if skipped:
        msg += f"\n⏩ Skipped {len(skipped)} existing members: {', '.join(skipped)}"

    return msg




@frappe.whitelist()
def create_loan_member():
    try:
        data = frappe.form_dict  # text fields
        files = frappe.request.files  # uploaded files

        # Get user and company
        user_doc = frappe.get_doc("User", frappe.session.user)
        try:
            emp_details = frappe.get_doc("Employee", {"user_id": user_doc.name})
            company = emp_details.company
        except:
            company = ""

        # Step 1: Create Loan Member doc (without files)
        doc = frappe.get_doc({
            "doctype": "Loan Member",
            "member_id": data.get("member_id"),
            "first_name": data.get("first_name"),
            "middle_name": data.get("middle_name"),
            "last_name": data.get("last_name"),
            "gender": data.get("gender"),
            "dob": data.get("dob"),
            "completed_age": data.get("completed_age"),
            "entry_age": data.get("entry_age"),
            "mobile_no": data.get("mobile_no"),
            "mobile_no_2":data.get("mobile_no_2"),
            "company": company,
            "state": data.get("state"),
            "country": data.get("country"),
            "city": data.get("city"),
            "pincode": data.get("pincode"),
            "status": data.get("status", "Open"),
            "occupation": data.get("occupation"),
            "address_line_2": data.get("address_line_2"),

            # Corrected mappings
            "group": data.get("group"),
            "email": data.get("email"),
            "address": data.get("address"),
            "aadhar": data.get("aadhar"),
            "pancard": data.get("pancard"),
            "voter_id":data.get("voter_id"),
            "address_doc_type": data.get("address_doc_type"),

            "cibil_score": data.get("cibil_score"),
            "cibil_date": data.get("cibil_date"),

            # Bank details
            "bank_details": data.get("bank_details"),
            "bank_name": data.get("bank_name"),
            "account_number": data.get("account_number"),
            "holder_name": data.get("holder_name"),
            "branch": data.get("branch"),
            "ifsc_code": data.get("ifsc_code"),
            "account_type": data.get("account_type"),
            "bank_address": data.get("bank_address"),

            "longitude": data.get("longitude"),
            "latitude": data.get("latitude"),
            "geo_location": data.get("geo_location"),

            "status": "Draft"
        })

        doc.insert(ignore_permissions=True)


        # Step 2: Handle file uploads
        for field in ["member_image", "aadhar_image", "pancard_image","address_image","home_image","voter_id_image",
                      "aadhar_image_back", "pancard_image_back","voter_id_image_back"]:
            if field in files:
                upload = files[field]
                if not upload or not upload.filename:
                    continue  
                file_doc = save_file(
                    fname=upload.filename,
                    content=upload.stream.read(),
                    dt="Loan Member",
                    dn=doc.name,
                    # dn=1,
                    is_private=0
                )
                doc.set(field, file_doc.file_url)

        # Step 3: Save updated doc with file URLs
        # doc.insert(ignore_permissions=True)
        doc.save(ignore_permissions=True)
        # frappe.db.commit()

        return {
            "status": "success",
            "status_code": 201,
            "msg": "Loan Member Created Successfully",
            "data":doc.as_dict()
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Loan Member Update API Error")
        return api_error(e)



update_fields = [
    "name",
    "first_name", "last_name", "middle_name", "gender", "dob",
    "completed_age", "entry_age", "mobile_no", "mobile_no_2",
    "state", "country",
    "city", "pincode", "status","occupation",
    "group",
    "email",
    "address", "address_line_2",
    "aadhar",
    "pancard",
    "aadhar_image",
    "pancard_image",
    "address_image",
    "aadhar_verified",
    "pancard_verified",
    "address_verified",
    "cibil_score",
    "cibil_date",
    "bank_name",
    "account_number",
    "holder_name",
    "branch",
    "ifsc_code",
    "account_type",
    "bank_address",
    "created_by",
    "member_id", "member_image", "company", "member_name",
    "address_doc_type", "home_image","voter_id","voter_id_image",
    "aadhar_image_back", "pancard_image_back","voter_id_image_back",
    "longitude", "latitude", "geo_location",
]

"""
	Get Loan Member List (with optional pagination, search & sorting)
"""

@frappe.whitelist()
def loan_member_list(page=1, page_size=10, search=None, sort_by="occupation", sort_order="asc",status=None, is_pagination=False, **kwargs):
    is_pagination = frappe.utils.sbool(is_pagination)  # convert "true"/"false"/1/0 into bool
    extra_params = {"search": search} if search else {}
    print("kwargs ...",kwargs)
    kwargs.pop('cmd', None)
    user = frappe.session.user

    # 🔹 Collect filters from kwargs (all query params except the defaults)
    filters = {}
    for k, v in kwargs.items():
        if k != 'is_group':
            if v not in [None, ""]:   # skip empty params
                filters[k] = v

    # 🔹 Handle is_group filter for Link field
    is_group = kwargs.get("is_group")
    if is_group is not None:
        is_group = frappe.utils.sbool(is_group)
        if is_group:
            # Members with a group assigned (group is not null/empty)
            # filters["group"] = ["!=", ""]
            employee_id = frappe.db.get_value("Employee", {"user_id": user}, "name")
            if not employee_id:
                return []
            
            if kwargs.get("group") == None or kwargs.get("group") == "":

                # Get loan groups for this employee
                groups = frappe.get_all(
                    "Loan Group Assignment",
                    filters={"employee": employee_id},
                    pluck="loan_group"
                )
                if not groups:
                    groups = []

                # 🔹 Collect filters from kwargs (all query params except the defaults)
                filters["group"] = ["in", groups]
        else:
            # Members without a group assigned (group is null/empty)
            filters["group"] = ["in", [None, ""]]
            filters["owner"] = frappe.session.user

    # filters["owner"] = frappe.session.user
    
    if status in ["Verified", "Pending"]:
        print("in if ...................", status)
        employee_id = frappe.db.get_value("Employee", {"user_id": user}, "name")
        if not employee_id:
            return []

        if kwargs.get("group") == None or kwargs.get("group") == "":
            # Get loan groups for this employee
            groups = frappe.get_all(
                "Loan Group Assignment",
                filters={"employee": employee_id},
                pluck="loan_group"
            )
            if not groups:
                groups = []

            # 🔹 Collect filters from kwargs (all query params except the defaults)
            filters["group"] = ["in", groups]
        filters["status"] = status

    if status == "Draft":
        filters["owner"] = frappe.session.user
        filters["status"] = status

    print("filters ...",filters)
    base_url = frappe.request.host_url.rstrip("/") + frappe.request.path

    return get_paginated_data(
        doctype="Loan Member",
        fields=update_fields,
        filters=filters,   # ✅ Now your filters will be applied
        search=search,
        sort_by=sort_by,
        sort_order=sort_order,
        page=int(page),
        page_size=int(page_size),
        search_fields=["member_name"],
        is_pagination=is_pagination,
        base_url=base_url,
        extra_params=extra_params,
        link_fields={"group": "group_name"},
        image_fields=['member_image','aadhar_image','pancard_image','address_image','home_image','voter_id_image',
                      "aadhar_image_back", "pancard_image_back","voter_id_image_back"]
    )



@frappe.whitelist()
def update_loan_member_api(name):
    try:
        data = frappe.form_dict   # text fields
        files = frappe.request.files  # uploaded files

        loan_member_id = name
        if not loan_member_id:
            return {"status": "error", "message": "Loan Member ID is required"}

        # # Fetch existing loan member doc
        doc = frappe.get_doc("Loan Member", loan_member_id)

        image_fields = ["member_image", "aadhar_image", "pancard_image","voter_id_image","home_image","address_image",
                             "aadhar_image_back", "pancard_image_back","voter_id_image_back"]

        # Update text fields
        for field in update_fields:
            if field not in image_fields:
                if data.get(field) is not None:
                    doc.set(field, data.get(field))

        # Handle file uploads OR existing path strings
        for field in image_fields:
            if field in files and files[field].filename:
                # If new file uploaded → save and replace
                upload = files[field]
                file_doc = save_file(
                    fname=upload.filename,
                    content=upload.stream.read(),
                    dt="Loan Member",
                    dn=doc.name,
                    is_private=0
                )
                doc.set(field, file_doc.file_url)
            elif data.get(field):
                # If frontend sends string (URL/path) → just set it
                doc.set(field, data.get(field))

        doc.set("status", "Draft")

        # Save changes
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {
            "status": "success",
            "status_code": 201,
            "msg": "Loan Member Updated Successfully",
            "data":doc.as_dict()
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Loan Member Update API Error")
        # return {"status": "error", "message": str(e)}
        return api_error(e)




@frappe.whitelist()
def update_status(docname, status):
    """Update verification_status of a Loan Member"""
    doc = frappe.get_doc("Loan Member", docname)
    doc.status = status
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return {"status": "success", "new_state": status}


@frappe.whitelist()
def update_loan_member(name):
    try:
        doc = frappe.get_doc("Loan Member", name)
        doc.set("status", "Pending")
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return {
            "status": "success",
            "status_code": 201,
            "msg": "Loan Member Sent for Verification Successfully"
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Loan Member Update API Error")
        return api_error(e)





@frappe.whitelist()
def loan_member_get(name):
    """
    Get Loan Member by name (primary key)
    Returns linked Loan Group name and full URLs for image fields
    """
    if not name:
        frappe.throw("Loan Member name is required")

    # Fetch Loan Member
    members = frappe.get_all(
        "Loan Member",
        filters={"name": name},
        fields=update_fields
    )

    if not members:
        return {}

    member = members[0]

    # 🔹 Get linked Loan Group name
    if member.get("group"):
        group_doc = frappe.get_value("Loan Group", member["group"], "group_name")
        member["group_group_name"] = group_doc or ""

    # 🔹 Full URL for image fields
    host_url = frappe.request.host_url.rstrip("/")  # e.g. http://localhost:8000
    image_fields = ['member_image','aadhar_image','pancard_image','address_image','home_image','voter_id_image',
                    "aadhar_image_back", "pancard_image_back","voter_id_image_back"]

    for f in image_fields:
        if member.get(f):
            member[f] = urljoin(host_url, member[f])

    return member


@frappe.whitelist()
def loan_member_list_as_per_group_assignment(page=1, page_size=10, search=None, sort_by="occupation", sort_order="asc", is_pagination=False, **kwargs):
    user = frappe.session.user

    if "Agent" in frappe.get_roles(user):
        # Get employee id
        employee_id = frappe.db.get_value("Employee", {"user_id": user}, "name")
        if not employee_id:
            return []

        # Get loan groups for this employee
        groups = frappe.get_all(
            "Loan Group Assignment",
            filters={"employee": employee_id},
            pluck="loan_group"
        )
        if not groups:
            groups = []

        is_pagination = frappe.utils.sbool(is_pagination)  # convert "true"/"false"/1/0 into bool
        extra_params = {"search": search} if search else {}
        del kwargs['cmd']

        # 🔹 Collect filters from kwargs (all query params except the defaults)
        filters = {}
        filters["group"] = ["in", groups]
        filters["status"] = ["in", "Verified"]  

        print("filters ...",filters)
        base_url = frappe.request.host_url.rstrip("/") + frappe.request.path

        return get_paginated_data(
            doctype="Loan Member",
            fields=update_fields,
            filters=filters,   # ✅ Now your filters will be applied
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
            page=int(page),
            page_size=int(page_size),
            search_fields=["member_name"],
            is_pagination=is_pagination,
            base_url=base_url,
            extra_params=extra_params,
            link_fields={"group": "group_name"},
            image_fields=['member_image','aadhar_image','pancard_image','address_image','home_image','voter_id_image']
        )
    else:
        return {}